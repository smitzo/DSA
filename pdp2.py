import openpyxl
from openpyxl.styles import Font, PatternFill
import customtkinter as ctk
from tkinter import filedialog, messagebox, StringVar, Toplevel, Label, PhotoImage
import os
import json
from PIL import Image, ImageTk
import threading
import tkinter as tk
import random


# Function to save settings
def save_settings():
    settings = {
        "theme_color": theme_color,
        "appearance_mode": ctk.get_appearance_mode()
    }
    with open("settings.json", "w") as f:
        json.dump(settings, f)

# Function to load settings
def load_settings():
    global theme_color
    try:
        with open("settings.json", "r") as f:
            settings = json.load(f)
            theme_color = settings.get("theme_color", "blue")
            ctk.set_appearance_mode(settings.get("appearance_mode", "Light"))
    except FileNotFoundError:
        theme_color = "blue"
        ctk.set_appearance_mode("Light")

def process_excel(file_path, column):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active

    min_row = 2
    max_row = sheet.max_row
    min_col = column  # Column for processing

    error_cells = []

    for row in range(min_row, max_row + 1):
        if sheet.cell(row=row, column=1).border.top.style:  # Check for border
            min_row = row + 1
            continue

        min_price = float('inf')
        min_price_row = None

        for r in range(min_row, max_row + 1):
            if sheet.cell(row=r, column=1).border.top.style:  # Check for border
                break

            price = sheet.cell(row=r, column=min_col).value
            if price < min_price:
                min_price = price
                min_price_row = r

        if min_price_row:
            for r in range(min_row, max_row + 1):
                if sheet.cell(row=r, column=1).border.top.style:  # Check for border
                    break

                price = sheet.cell(row=r, column=min_col).value
                if price == min_price:
                    sheet.cell(row=r, column=min_col).font = Font(bold=True)
                else:
                    sheet.cell(row=r, column=min_col).fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                    error_cells.append((r, min_col))

    processed_file_path = "processed_" + file_path.split("/")[-1]
    wb.save(processed_file_path)
    return processed_file_path, error_cells


def upload_file():
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if file_path:
        file_label.configure(text=os.path.basename(file_path), text_color=theme_color)
        global uploaded_file
        uploaded_file = file_path
        result_box.insert(ctk.END, f"File uploaded: {os.path.basename(file_path)}\n")

def process_file():
    if uploaded_file:
        result_box.insert(ctk.END, "Processing...\n")
        threading.Thread(target=process_file_thread).start()
    else:
        result_label.configure(text="Please upload a file first.", text_color="red")
        result_box.insert(ctk.END, "Please upload a file first.\n")

def process_file_thread():
    processed_file, error_cells = process_excel(uploaded_file, ord(column_var.get().upper()) - 64)
    result_label.configure(text=f"File processed successfully: {processed_file}", text_color=theme_color)
    result_box.insert(ctk.END, f"File processed successfully: {processed_file}\n")
    if error_cells:
        result_box.insert(ctk.END, f"Error cells: {error_cells}\n")
    else:
        result_box.insert(ctk.END, "No errors found.\n")


import pandas as pd
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font

def highlight_cells():
    global uploaded_file
    if uploaded_file:
        # Read the Excel file, skipping the first two rows
        df = pd.read_excel(uploaded_file, skiprows=2)

        # Rename columns for easier access
        df.columns = ['Row Labels', 'Average of Output Unit Cost', 'Sum of Output Qty']

        # Load workbook to check for bold text
        wb = load_workbook(uploaded_file)
        sheet = wb.active

        # Variable to hold the current main product
        current_main_product = None

        # Iterate over the DataFrame to check for bold values in the Excel file
        for idx, row in df.iterrows():
            # Adjusting index for Excel row (consider skipping 2 rows)
            excel_row_num = idx + 3  # Excel row starts from 3 since we skipped 2 rows
            
            # Check if the cell in 'Row Labels' column has bold text
            if sheet[f'A{excel_row_num}'].font.bold:
                # If bold, it's a main product
                current_main_product = row['Row Labels']
            # Assign the current main product (whether bold or not)
            df.at[idx, 'Main Product'] = current_main_product

        # Log DataFrame after assigning Main Product labels
        print("DataFrame after assigning Main Product labels:")
        print(df.head(20))

        # Remove rows where 'Row Labels' are None or empty
        df_cleaned = df[df['Row Labels'].notna() & (df['Row Labels'] != '')].copy()

        # Log DataFrame after cleaning
        print("DataFrame after cleaning (removing rows with missing labels):")
        print(df_cleaned.head(20))

        # Group by 'Main Product' and calculate thresholds
        def calculate_thresholds(group):
            cost_min = group['Average of Output Unit Cost'].min()
            cost_max = group['Average of Output Unit Cost'].max()
            cost_mean = group['Average of Output Unit Cost'].mean()
            cost_75th = group['Average of Output Unit Cost'].quantile(0.75)
            qty_max = group['Sum of Output Qty'].max()

            return cost_min, cost_max, cost_mean, cost_75th, qty_max

        # Group without including grouping column
        thresholds = df_cleaned.groupby('Main Product', group_keys=False).apply(calculate_thresholds).to_dict()

        # Log thresholds
        print("Thresholds calculated for each Main Product group:")
        print(thresholds)

        # Function to assign color based on dynamic thresholds
        def determine_color_code(row):
            product = row['Main Product']
            cost = row['Average of Output Unit Cost']
            qty = row['Sum of Output Qty']

            # Fetch thresholds for this product
            if product in thresholds:
                cost_min, cost_max, cost_mean, cost_75th, qty_max = thresholds[product]

                # Apply color coding logic
                if cost > cost_mean and qty > cost_75th:
                    return 'Red'  # High cost, high quantity
                elif cost_min < cost <= cost_mean and qty > cost_75th:
                    return 'Yellow'  # Medium cost, significant quantity
                elif cost <= cost_min and qty == qty_max:
                    return 'Green'  # Lowest cost, most quantity
                elif cost > cost_min and cost <= 1.2 * cost_min and qty > 0.2 * qty_max:
                    return 'Blue'  # Special condition
                else:
                    return 'No Color'
            else:
                return 'No Color'

        # Apply color code function to the DataFrame
        df_cleaned['Color Code'] = df_cleaned.apply(determine_color_code, axis=1)

        # Log DataFrame after applying color coding
        print("DataFrame after applying color coding:")
        print(df_cleaned.head(20))

        # Save the DataFrame back into an Excel file
        output_file = 'highlighted_output.xlsx'
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            df_cleaned.to_excel(writer, index=False)

        print(f"Excel file saved as {output_file}")
    
    else:
        result_label.configure(text="Please upload a file first.", text_color="red")


# def highlight_cells():
#     global uploaded_file
#     if uploaded_file:
#         # Read the Excel file, skipping the first two rows
#         df = pd.read_excel(uploaded_file, skiprows=2)

#         # Rename columns for easier access
#         df.columns = ['Row Labels', 'Average of Output Unit Cost', 'Sum of Output Qty']

#         # Load workbook to check for bold text
#         wb = load_workbook(uploaded_file)
#         sheet = wb.active

#         # Iterate over the DataFrame to check for bold values
#         for idx, row in df.iterrows():
#             # Check if the corresponding cell in Excel has bold text
#             if sheet[f'A{idx+3}'].font.bold:  # Adjust for the row skip (row 3 is the actual first data row)
#                 # Replace the bold value with a placeholder (e.g., '-')
#                 df.at[idx, 'Row Labels'] = '-'
#             if sheet[f'B{idx+3}'].font.bold:
#                 df.at[idx, 'Average of Output Unit Cost'] = '-'
#             if sheet[f'C{idx+3}'].font.bold:
#                 df.at[idx, 'Sum of Output Qty'] = '-'

#         # Filter out rows with placeholder '-'
#         df_cleaned = df[df['Row Labels'] != '-'].copy()
        
#         # Group by main product (based on cleaned data)
#         df_cleaned['Main Product'] = df_cleaned['Row Labels'].str.extract(r'([A-Za-z]+)').fillna('Unknown')
        
#         # For each product group, calculate the dynamic thresholds
#         def calculate_thresholds(group):
#             cost_min = group['Average of Output Unit Cost'].min()
#             cost_max = group['Average of Output Unit Cost'].max()
#             cost_mean = group['Average of Output Unit Cost'].mean()
#             cost_75th = group['Average of Output Unit Cost'].quantile(0.75)
#             qty_max = group['Sum of Output Qty'].max()

#             return cost_min, cost_max, cost_mean, cost_75th, qty_max

#         thresholds = df_cleaned.groupby('Main Product').apply(calculate_thresholds).to_dict()

#         # Function to assign color based on dynamic thresholds
#         def determine_color_code(row):
#             product = row['Main Product']
#             cost = row['Average of Output Unit Cost']
#             qty = row['Sum of Output Qty']

#             # Fetch thresholds for this product
#             cost_min, cost_max, cost_mean, cost_75th, qty_max = thresholds.get(product, (0, 0, 0, 0, 0))

#             # Apply color coding logic
#             if cost > cost_mean and qty > cost_75th:
#                 return 'Red'  # High cost, high quantity
#             elif cost_min < cost <= cost_mean and qty > cost_75th:
#                 return 'Yellow'  # Medium cost, significant quantity
#             elif cost <= cost_min and qty == qty_max:
#                 return 'Green'  # Lowest cost, most quantity
#             elif cost > cost_min and cost <= 1.2 * cost_min and qty > 0.2 * qty_max:
#                 return 'Blue'  # Special condition
#             else:
#                 return 'No Color'

#         # Apply color code function to the DataFrame
#         df_cleaned['Color Code'] = df_cleaned.apply(determine_color_code, axis=1)

#         # Display the top 5 rows of the new DataFrame
#         print("Top 5 rows of the new DataFrame with Color Code:")
#         print(df_cleaned.head(20))  # Show the top 20 rows for better visibility

#         # Save the DataFrame back into an Excel file
#         output_file = 'highlighted_output.xlsx'
#         with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
#             df_cleaned.to_excel(writer, index=False)

#         print(f"Excel file saved as {output_file}")
    
#     else:
#         result_label.configure(text="Please upload a file first.", text_color="red")

# def highlight_cells():
#     global uploaded_file
#     if uploaded_file:
#         # Read the Excel file into a DataFrame, skipping the first two rows
#         df = pd.read_excel(uploaded_file, skiprows=2)

#         # Rename the columns for easier access
#         df.columns = ['Row Labels', 'Average of Output Unit Cost', 'Sum of Output Qty']

#         # Calculate minimum and maximum costs for color coding
#         min_cost = df['Average of Output Unit Cost'].min()
#         max_cost = df['Average of Output Unit Cost'].max()
#         mean_cost = df['Average of Output Unit Cost'].mean()
#         max_qty = df['Sum of Output Qty'].max()  # Get the maximum quantity once

#         # Define thresholds for color coding
#         print(f"Min Cost: {min_cost}, Max Cost: {max_cost}, Mean Cost: {mean_cost}")
        
#         # Create the Color Code column based on the specified conditions
#         def determine_color_code(row):
#             cost = row['Average of Output Unit Cost']
#             qty = row['Sum of Output Qty']

#             print(f"Cost: {cost}, Quantity: {qty}")  # Debugging output

#             # Define thresholds for quantity
#             high_qty_threshold = 5000
#             medium_qty_threshold = 3000
#             special_qty_threshold = 2000

#             # Determine color code based on conditions
#             if cost > mean_cost and qty > high_qty_threshold:
#                 return 'Red'
#             elif min_cost < cost <= mean_cost and qty > medium_qty_threshold:
#                 return 'Yellow'
#             elif cost <= min_cost and qty == max_qty:
#                 return 'Green'
#             elif cost > min_cost and cost <= 1.1 * min_cost and qty > special_qty_threshold:
#                 return 'Blue'
#             else:
#                 return 'No Color'

#         # Apply the function to create the new column
#         df['Color Code'] = df.apply(determine_color_code, axis=1)

#         # Display the top 5 rows of the new DataFrame
#         print("Top 5 rows of the new DataFrame with Color Code:")
#         print(df.head(20))  # Show the top 20 rows for better visibility

#     else:
#         result_label.configure(text="Please upload a file first.", text_color="red")


def download_file():
    if uploaded_file:
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if file_path:
            processed_file, _ = process_excel(uploaded_file, ord(column_var.get().upper()) - 64)
            os.rename(processed_file, file_path)
            result_box.insert(ctk.END, f"File saved to: {file_path}\n")
    else:
        result_label.configure(text="Please upload a file first.", text_color="red")
        result_box.insert(ctk.END, "Please upload a file first.\n")

def reset():
    file_label.configure(text="No file uploaded", text_color=theme_color)
    result_label.configure(text="")
    global uploaded_file
    uploaded_file = None
    result_box.insert(ctk.END, "Reset completed.\n")

def open_settings():
    settings_frame.grid(row=0, column=0, padx=10, pady=10, sticky="nsew")
    main_frame.grid_forget()

def close_settings():
    settings_frame.grid_forget()
    main_frame.grid(row=0, column=0, padx=10, pady=10, sticky="nsew")

def switch_mode():
    if ctk.get_appearance_mode() == "Light":
        ctk.set_appearance_mode("Dark")
        result_label.configure(text_color="white")
        result_box.configure(text_color="white")
    else:
        ctk.set_appearance_mode("Light")
        result_label.configure(text_color="black")
        result_box.configure(text_color="black")
    save_settings()

def change_theme_color(color):
    global theme_color
    if color == "transparent":
        theme_color = "blue"  # Default color
    else:
        theme_color = color
    file_label.configure(text_color=theme_color)
    result_label.configure(text_color=theme_color)
    settings_button.configure(fg_color=theme_color)
    reset_button.configure(fg_color=theme_color)
    back_button.configure(fg_color=theme_color)
    switch_button.configure(fg_color=theme_color)
    about_button.configure(fg_color=theme_color)
    for widget in main_frame.winfo_children():
        if isinstance(widget, ctk.CTkButton):
            widget.configure(fg_color=theme_color)
    bottom_line.configure(fg_color=theme_color)
    settings_bottom_line.configure(fg_color=theme_color)
    save_settings()

def show_about():
    messagebox.showinfo("About", 
        "Pdp Processor v1.0 \nDeveloped by Jiren Pandya \nOther Contributor:\nApplied by:  \nThis software is the property of . \nThis software helps in analyzing Excel sheets. \nIn upcoming updates, we will provide time graph analysis, bug fixes, and additional features. \nThe user ID is 0 and the password is 0. \nThe password and user ID are unchangeable. \nFor any help regarding this software, please call +91 8401751355 (Jiren Pandya) or email at pandyajiren15@gmail.com.")
def authenticate_user():
    def check_credentials():
        user_id = user_id_entry.get()
        password = password_entry.get()
        if user_id == "0" and password == "0":
            auth_window.destroy()
            app.deiconify()  # Show the main app window
        elif user_id == "dev" and password == "000000":
            auth_window.destroy()
            app.deiconify()  # Show the main app window
        else:
            messagebox.showerror("Error", "Invalid credentials")

    auth_window = Toplevel(app)
    auth_window.title("Authentication")
    auth_window.geometry("955x645")

# Disable resizing the window
    auth_window.resizable(False, False)
    

    Label(auth_window, text="User ID:").pack(pady=5)
    user_id_entry = ctk.CTkEntry(auth_window)
    user_id_entry.pack(pady=5)

    Label(auth_window, text="Password:").pack(pady=5)
    password_entry = ctk.CTkEntry(auth_window, show="*")
    password_entry.pack(pady=5)

    ctk.CTkButton(auth_window, text="Login", command=check_credentials).pack(pady=10)

app = ctk.CTk()
app.title("Paint Details Processor")
app.geometry("955x645")

# Disable resizing the window
app.resizable(False, False)
app.withdraw()  # Hide the main app window initially

uploaded_file = None
theme_color = "blue"

# Load settings
load_settings()

# Authenticate user
authenticate_user()

# Main Frame
main_frame = ctk.CTkFrame(app)
main_frame.grid(row=0, column=0, padx=10, pady=10, sticky="nsew")

# Upload button
upload_button = ctk.CTkButton(main_frame, text="Upload Excel Sheet", command=upload_file, fg_color=theme_color, width=130, height=50)
upload_button.grid(row=0, column=0, padx=10, pady=10)

# Process button
process_button = ctk.CTkButton(main_frame, text="   Process File  ", command=process_file, fg_color=theme_color, width=130, height=50)
process_button.grid(row=1, column=0, padx=10, pady=10)

# Highlight button
highlight_button = ctk.CTkButton(main_frame, text="Highlight Cells", command=highlight_cells, fg_color=theme_color, width=130, height=50)
highlight_button.grid(row=2, column=0, padx=10, pady=10)

# Download button
download_button = ctk.CTkButton(main_frame, text="Download File", command=download_file, fg_color=theme_color, width=130, height=50)
download_button.grid(row=3, column=0, padx=10, pady=10)

# File label
file_label = ctk.CTkLabel(main_frame, text="No file uploaded", text_color=theme_color)
file_label.grid(row=4, column=0, padx=10, pady=10)

# Result label
result_label = ctk.CTkLabel(main_frame, text="", text_color="black")
result_label.grid(row=5, column=0, padx=10, pady=10)

# Result# Result box
border_frame = ctk.CTkFrame(main_frame, corner_radius=10, border_width=2, border_color="grey")
border_frame.grid(row=0, column=1, rowspan=6, padx=10, pady=10)

# Create a Textbox inside the frame
result_box = ctk.CTkTextbox(border_frame, height=500, width=580)
result_box.pack(fill="both", expand=True, padx=5, pady=5) 
# Settings button with icon
settings_button = ctk.CTkButton(main_frame, text="⚙️", command=open_settings, fg_color=theme_color, width=50, height=50)
settings_button.grid(row=0, column=2, padx=10, pady=10)

# Reset button with icon
reset_button = ctk.CTkButton(main_frame, text="⟳", command=reset, fg_color=theme_color, width=50, height=50)
reset_button.grid(row=0, column=3, padx=10, pady=10)

# Bottom line
bottom_line = ctk.CTkFrame(main_frame, height=2, fg_color=theme_color)
bottom_line.grid(row=6, column=0, columnspan=4, padx=10, pady=5, sticky="ew")

# Logo

# Load and resize the image
#logo_image = Image.open("C:/Users/SMIT/OneDrive/Desktop/jiren/PP/pdp.png")
#logo_image = logo_image.resize((100, 100), Image.LANCZOS)  # Adjust size as needed

# Ensure that the image has an alpha channel for transparency
#logo_image = logo_image.convert("RGBA")

# Convert to a format suitable for tkinter
#logo_photo = ImageTk.PhotoImage(logo_image)

# Create a label and set the image
#logo_label = tk.Label(main_frame, image=logo_photo, bg="white")  # Use any solid background if needed
#logo_label.image = logo_photo  # Keep a reference to avoid garbage collection
#logo_label.grid(row=6, column=3, padx=10, pady=5, sticky="e")
# Settings Frame
settings_frame = ctk.CTkFrame(app)

# Back button
back_button = ctk.CTkButton(settings_frame, text="<", command=close_settings, fg_color=theme_color,width=50, height=50)
back_button.grid(row=0, column=0, padx=10, pady=10, sticky="nw")

# Switch mode button
switch_button = ctk.CTkSwitch(settings_frame, text="Dark/Light Mode", command=switch_mode)
switch_button.grid(row=1, column=0, padx=10, pady=10, sticky="nw")

# Theme color options
theme_label = ctk.CTkLabel(settings_frame, text="Select Theme Color")
theme_label.grid(row=2, column=0, padx=10, pady=10, sticky="nw")

colors = [
    "black", "darkblue", "darkred", "dark green", "darkslategrey", "SlateGray",
    "Indigo", "DarkViolet", "SaddleBrown", "olive", "tan", "blue", "dodger blue",
    "#007BFF", "cyan4", "green", "lime green", "orange", "red", "purple",
    "deep pink", "transparent"
     ]
color_buttons = ctk.CTkFrame(settings_frame)
color_buttons.grid(row=3, column=0, padx=10, pady=10, sticky="nw")

for color in colors:
    color_button = ctk.CTkButton(color_buttons, text="✔" if color == theme_color else "", fg_color=color if color != "transparent" else "white", width=30, height=30, command=lambda c=color: change_theme_color(c))
    color_button.pack(side="left", padx=5)

# About button
about_button = ctk.CTkButton(settings_frame, text="About", command=show_about, fg_color=theme_color,width=130, height=50)
about_button.grid(row=4, column=0, padx=10, pady=10, sticky="nw")

# Default column for processing
column_var = StringVar(value="D")
column_label = ctk.CTkLabel(settings_frame, text="Default Column for Processing")
column_label.grid(row=5, column=0, padx=10, pady=10, sticky="nw")
column_entry = ctk.CTkEntry(settings_frame, textvariable=column_var)
column_entry.grid(row=6, column=0, padx=10, pady=10, sticky="nw")


#pimage = Image.open("C:/Users/SMIT/OneDrive/Desktop/jiren/PP/b.png")
#pimage = pimage.resize((100, 100), Image.LANCZOS)  # Adjust size as needed

# Ensure that the image has an alpha channel for transparency
#pimage = pimage.convert("RGBA")

# Convert to a format suitable for tkinter
#pphoto = ImageTk.PhotoImage(pimage)

# Create a label and set the image
#plabel = tk.Label(settings_frame, image=pphoto, bg="white")  # Use any solid background if needed
#plabel.image = pphoto  # Keep a reference to avoid garbage collection
#plabel.grid(row=6, column=0, padx=10, pady=5, sticky="e")
# Bottom line for settings frame
settings_bottom_line = ctk.CTkFrame(settings_frame, height=2, fg_color=theme_color)
settings_bottom_line.grid(row=7, column=0, columnspan=2, padx=10, pady=5, sticky="ew")

app.mainloop()
